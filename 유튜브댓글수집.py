'''
    try - except : indexerror 발생하게되면 종료하도록함. 한 페이지에 보여지는 댓글을 20개 기준으로 하기 때문에
    스크롤 끝 지점의 댓글 수가 20개 미만인 경우 범위 오류 발생.
    수집한 댓글 엑셀파일로 저장. -> 테스트용이라 sheet를 하나로 고정.
'''

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import os
import openpyxl


browser = webdriver.Chrome("./chromedriver.exe")
browser.get("https://www.youtube.com/watch?v=IdHk0U9fL_c")
time.sleep(8)

browser.find_element_by_css_selector("html").send_keys(Keys.PAGE_DOWN)
time.sleep(8)

title = browser.find_elements_by_css_selector('#content-text')
cnt = 0

if os.path.exists("./댓글수집.xlsx") == False:
    openpyxl.Workbook().save("./댓글수집.xlsx")

comments = openpyxl.load_workbook("./댓글수집.xlsx")
sheet = comments.create_sheet()
sheet.title = "유튜브"
if "Sheet" in comments.sheetnames:
    comments.remove(comments["Sheet"])
row_num = 1
while True :
    try:
        sheet.cell(row=row_num, column=1).value = title[cnt].text
        row_num +=1
        comments.save("./댓글수집.xlsx")
        print(title[cnt].text)
        cnt += 1
        if cnt % 20 == 0:
            browser.find_element_by_css_selector("html").send_keys(Keys.END)
            time.sleep(8)
            try:
                title = browser.find_elements_by_css_selector("#content-text")
            except IndexError:
                print("종료")
                break
    except IndexError:
        print("종료")
        break